"""XLSX grid text extraction via openpyxl (read-only, values only)."""

from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from app.parsers._versions import package_version
from app.parsers.base import DocumentParser, ParserResult
from app.services.text_extraction_service import normalize_extension


class XlsxParser(DocumentParser):
    PARSER_NAME = "xlsx_parser"

    def __init__(self) -> None:
        self._version = package_version("openpyxl")

    def supports(self, extension: str, mime_type: str | None) -> bool:
        return normalize_extension(extension) == "xlsx"

    def parse_bytes(
        self, content: bytes, filename: str, extension: str
    ) -> ParserResult:
        try:
            wb = load_workbook(
                BytesIO(content),
                read_only=True,
                data_only=True,
            )
        except Exception:
            return ParserResult(
                success=False,
                extracted_text=None,
                text_length=0,
                parser_name=self.PARSER_NAME,
                parser_version=self._version,
                metadata={},
                error_code="PARSING_FAILED",
                error_message="Document parsing failed",
            )

        sheet_names: list[str] = list(wb.sheetnames)
        sections: list[str] = []
        non_empty_rows = 0

        try:
            for name in sheet_names:
                ws = wb[name]
                rows_out: list[str] = []
                for row in ws.iter_rows(values_only=True):
                    if row is None:
                        continue
                    cells = [
                        "" if v is None else str(v).replace("\n", " ").replace("\t", " ").strip()
                        for v in row
                    ]
                    if not any(cells):
                        continue
                    non_empty_rows += 1
                    rows_out.append("\t".join(cells))
                if rows_out:
                    sections.append(f"--- sheet: {name} ---\n" + "\n".join(rows_out))
        finally:
            try:
                wb.close()
            except Exception:
                pass

        text = "\n\n".join(sections).strip()
        return ParserResult(
            success=True,
            extracted_text=text,
            text_length=len(text),
            parser_name=self.PARSER_NAME,
            parser_version=self._version,
            metadata={
                "sheet_count": len(sheet_names),
                "sheet_names": sheet_names,
                "non_empty_rows": non_empty_rows,
            },
        )


__all__ = ["XlsxParser"]
